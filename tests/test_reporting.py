from openpyxl import load_workbook
from airfryer_rankings.reporting import write_dashboard, write_workbook


def sample_ranked():
    return [{
        "rank": 1, "recipe_id": "a", "title": "Air Fryer Chicken Bites", "source": "x.com",
        "combined_sources": "x.com", "url": "https://x.com/a", "rating": 4.9, "rating_count": 500,
        "source_bias": 0.02, "adjusted_rating": 4.88, "posterior_mean": 4.85,
        "uncertainty_penalty": 0.1, "hierarchical_score": 4.75, "evidence_confidence": 1.0,
        "evidence_status": "verified", "author": "A", "categories": "Chicken | Snacks",
        "duplicate_group_id": "", "duplicate_confidence": 0.0, "last_seen_at": "2026-08-18T20:00:00+00:00",
        "rating_change": 0.0, "review_count_change": 2, "review_velocity_per_day": 48.0, "previous_rank": 2, "movement": 1,
    }]


def test_workbook_contains_research_sheets(tmp_path):
    path = tmp_path / "rankings.xlsx"
    write_workbook(path, sample_ranked(), [], [], [], [], [], {"formula": "test"})
    wb = load_workbook(path, read_only=True)
    expected = {"Top 50", "All Rankings", "Source Coverage", "Source Reliability", "Rating History", "Rating Trends", "New Entrants", "Biggest Movers", "QA Anomalies", "Duplicate Groups", "Methodology", "Chicken", "Potatoes", "Vegetables", "Desserts", "Beef", "Pork", "Seafood", "Breakfast", "Snacks"}
    assert expected.issubset(set(wb.sheetnames))


def test_dashboard_is_searchable(tmp_path):
    write_dashboard(tmp_path, "2026-08-18T20:00:00+00:00", sample_ranked(), [], [], {"formula": "test"}, 40)
    html = (tmp_path / "index.html").read_text()
    data = (tmp_path / "data.json").read_text()
    assert 'id="search"' in html
    assert "Air Fryer Chicken Bites" in data


def test_rating_trends_sheet_contains_growth_chart(tmp_path):
    path = tmp_path / "rankings.xlsx"
    observations = [
        {"recipe_id": "a", "timestamp": "2026-08-18T19:00:00+00:00", "rating_count": 490},
        {"recipe_id": "a", "timestamp": "2026-08-18T20:00:00+00:00", "rating_count": 500},
    ]
    write_workbook(path, sample_ranked(), [], [], observations, [], [], {"formula": "test"})
    wb = load_workbook(path)
    assert len(wb["Rating Trends"]._charts) == 1
