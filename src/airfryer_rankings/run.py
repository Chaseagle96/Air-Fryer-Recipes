from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .core import (
    bayesian_rank,
    discover_domain,
    load_sources,
    load_state,
    merge_observations,
    now_iso,
    save_state,
)


def _style_workbook(path: str) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.alignment = Alignment(vertical="center")
        for col_idx, col in enumerate(ws.columns, 1):
            values = [str(c.value) if c.value is not None else "" for c in col[:100]]
            width = min(45, max(10, max((len(v) for v in values), default=10) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--sources", default="config/sources.yaml")
    ap.add_argument("--state", default="data/state.json")
    ap.add_argument("--max-urls", type=int, default=None)
    ap.add_argument("--stale-days", type=int, default=14)
    args = ap.parse_args()

    Path("data").mkdir(exist_ok=True)
    Path("output").mkdir(exist_ok=True)

    run_at = now_iso()
    state = load_state(args.state)
    coverage = []
    observed = 0

    for cfg in load_sources(args.sources):
        try:
            rows, result = discover_domain(cfg, global_max_urls=args.max_urls)
            merge_observations(state, rows, run_at)
            observed += len(rows)
            coverage.append(result)
        except Exception as exc:
            coverage.append(
                {
                    "source": cfg.domain,
                    "candidate_urls": 0,
                    "verified_recipes": 0,
                    "sitemap_docs": 0,
                    "elapsed_seconds": 0,
                    "status": f"error: {type(exc).__name__}: {str(exc)[:120]}",
                }
            )

    state.setdefault("source_history", []).append({"run_at": run_at, "coverage": coverage})
    if len(state["source_history"]) > 168:
        del state["source_history"][:-168]

    ranked, method = bayesian_rank(state, stale_days=args.stale_days)
    save_state(args.state, state)

    df = pd.DataFrame(ranked)
    cov = pd.DataFrame(coverage)
    movers = df[df["movement"].notna()].copy() if not df.empty and "movement" in df else pd.DataFrame()
    if not movers.empty:
        movers["abs_movement"] = movers["movement"].abs()
        movers = movers.sort_values(["abs_movement", "rating_count"], ascending=[False, False]).drop(columns=["abs_movement"])

    if not df.empty:
        df.to_csv("output/leaderboard.csv", index=False)
        df.head(50).to_csv("output/top50.csv", index=False)
    else:
        Path("output/leaderboard.csv").write_text("")
        Path("output/top50.csv").write_text("")

    method_row = {
        "generated_at": run_at,
        "observations_this_run": observed,
        "ranked_recipes": len(ranked),
        **method,
        "formula": "WR=(v/(v+m))*R + (m/(v+m))*C",
        "prior_definition": "C=sqrt(review_count)-weighted mean rating; m=max(50, 60th percentile review count)",
        "dedupe_definition": "Exact normalized title + exact ingredient signature; duplicate listings combine ratings by review-count weighting",
    }

    xlsx_path = "output/air_fryer_rankings.xlsx"
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.head(50).to_excel(writer, index=False, sheet_name="Top 50")
        df.to_excel(writer, index=False, sheet_name="All Rankings")
        cov.to_excel(writer, index=False, sheet_name="Source Coverage")
        movers.head(100).to_excel(writer, index=False, sheet_name="Movers")
        pd.DataFrame([method_row]).to_excel(writer, index=False, sheet_name="Methodology")
    _style_workbook(xlsx_path)

    summary = {
        **method_row,
        "source_count": len(coverage),
        "sources_ok": sum(1 for x in coverage if x["status"] == "ok"),
        "top10": ranked[:10],
        "coverage": coverage,
    }
    Path("output/summary.json").write_text(json.dumps(summary, indent=2) + "\n")


if __name__ == "__main__":
    main()
