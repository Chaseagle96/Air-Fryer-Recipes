from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .dashboard import write_dashboard

CATEGORY_SHEETS = ["Chicken", "Potatoes", "Vegetables", "Desserts", "Beef", "Pork", "Seafood", "Breakfast", "Snacks"]


def _df(rows) -> pd.DataFrame:
    if isinstance(rows, pd.DataFrame):
        return rows.copy()
    return pd.DataFrame(list(rows or []))


def _style_workbook(path: str) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row >= 1:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for col_idx, col in enumerate(ws.columns, 1):
            values = [str(c.value) if c.value is not None else "" for c in col[:100]]
            width = min(48, max(10, max((len(v) for v in values), default=10) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    if "Rating Trends" in wb.sheetnames:
        ws = wb["Rating Trends"]
        if ws.max_row >= 3 and ws.max_column >= 2:
            chart = LineChart()
            chart.title = "Rating-count growth for current Top 10"
            chart.y_axis.title = "Rating count"
            chart.x_axis.title = "Observation timestamp"
            data = Reference(ws, min_col=2, max_col=ws.max_column, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 12
            chart.width = 24
            ws.add_chart(chart, "A18")
    wb.save(path)


def write_csv_outputs(
    output_dir: str | Path,
    ranked: list[dict],
    coverage: list[dict],
    reliability: list[dict],
    anomalies: list[dict],
    *,
    source_health: list[dict] | None = None,
    robustness: list[dict] | None = None,
    dedupe_benchmark: list[dict] | None = None,
) -> None:
    output = Path(output_dir)
    output.mkdir(parents=True, exist_ok=True)
    df = _df(ranked)
    if df.empty:
        (output / "leaderboard.csv").write_text("")
        (output / "top50.csv").write_text("")
    else:
        df.to_csv(output / "leaderboard.csv", index=False)
        df.head(50).to_csv(output / "top50.csv", index=False)
    _df(coverage).to_csv(output / "source_coverage.csv", index=False)
    _df(reliability).to_csv(output / "source_reliability.csv", index=False)
    _df(anomalies).to_csv(output / "anomalies.csv", index=False)
    _df(source_health).to_csv(output / "source_health.csv", index=False)
    _df(robustness).to_csv(output / "ranking_robustness.csv", index=False)
    _df(dedupe_benchmark).to_csv(output / "dedupe_benchmark.csv", index=False)


def write_workbook(
    path: str | Path,
    ranked: list[dict],
    coverage: list[dict],
    reliability: list[dict],
    recent_observations: list[dict],
    anomalies: list[dict],
    duplicate_groups: list[dict],
    methodology: dict,
    *,
    source_health: list[dict] | None = None,
    uncertainty_calibration: list[dict] | None = None,
    robustness: list[dict] | None = None,
    dedupe_benchmark: list[dict] | None = None,
) -> None:
    path = str(path)
    df = _df(ranked)
    cov = _df(coverage)
    rel = _df(reliability)
    health = _df(source_health)
    obs = _df(recent_observations)
    anom = _df(anomalies)
    dup = _df(duplicate_groups)
    calibration = _df(uncertainty_calibration)
    robust = _df(robustness)
    benchmark = _df(dedupe_benchmark)
    trend_chart = pd.DataFrame()
    if not obs.empty and not df.empty and {"recipe_id", "timestamp", "rating_count"}.issubset(obs.columns):
        top_ids = [str(x) for x in df.head(10)["recipe_id"].tolist()] if "recipe_id" in df else []
        trend_source = obs[obs["recipe_id"].astype(str).isin(top_ids)].copy()
        if not trend_source.empty:
            labels = dict(zip(df["recipe_id"].astype(str), df["title"].astype(str))) if {"recipe_id", "title"}.issubset(df.columns) else {}
            trend_source["label"] = trend_source["recipe_id"].astype(str).map(labels).fillna(trend_source["recipe_id"].astype(str))
            trend_source["timestamp"] = pd.to_datetime(trend_source["timestamp"], errors="coerce", utc=True)
            trend_source = trend_source.dropna(subset=["timestamp"])
            trend_chart = trend_source.pivot_table(index="timestamp", columns="label", values="rating_count", aggfunc="last").sort_index().reset_index()
            if "timestamp" in trend_chart:
                trend_chart["timestamp"] = trend_chart["timestamp"].dt.tz_localize(None)

    movers = pd.DataFrame()
    entrants = pd.DataFrame()
    provenance = pd.DataFrame()
    if not df.empty:
        if "movement" in df:
            movers = df[df["movement"].notna()].copy()
            if not movers.empty:
                movers["abs_movement"] = movers["movement"].abs()
                movers = movers.sort_values(["abs_movement", "rating_count"], ascending=[False, False]).drop(columns=["abs_movement"])
        if "previous_rank" in df:
            entrants = df[df["previous_rank"].isna()].head(100).copy()
        provenance_columns = [
            x for x in (
                "rank", "title", "source", "rating", "rating_count", "category_expected_rating", "source_bias",
                "adjusted_rating", "posterior_mean", "uncertainty_penalty", "uncertainty_method",
                "evidence_penalty", "evidence_grade", "hierarchical_score", "rank_confidence",
                "rank_range_low", "rank_range_high", "rank_provenance", "url",
            ) if x in df.columns
        ]
        provenance = df[provenance_columns].head(200).copy()

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.head(50).to_excel(writer, index=False, sheet_name="Top 50")
        df.to_excel(writer, index=False, sheet_name="All Rankings")
        provenance.to_excel(writer, index=False, sheet_name="Rank Explainability")
        cov.to_excel(writer, index=False, sheet_name="Source Coverage")
        health.to_excel(writer, index=False, sheet_name="Source Health")
        rel.to_excel(writer, index=False, sheet_name="Source Reliability")
        obs.to_excel(writer, index=False, sheet_name="Rating History")
        trend_chart.to_excel(writer, index=False, sheet_name="Rating Trends")
        calibration.to_excel(writer, index=False, sheet_name="Uncertainty Calibration")
        robust.to_excel(writer, index=False, sheet_name="Rank Robustness")
        entrants.to_excel(writer, index=False, sheet_name="New Entrants")
        movers.head(200).to_excel(writer, index=False, sheet_name="Biggest Movers")
        anom.to_excel(writer, index=False, sheet_name="QA Anomalies")
        dup.to_excel(writer, index=False, sheet_name="Duplicate Groups")
        benchmark.to_excel(writer, index=False, sheet_name="Dedupe Benchmark")
        pd.DataFrame([methodology]).to_excel(writer, index=False, sheet_name="Methodology")
        for category in CATEGORY_SHEETS:
            if df.empty or "categories" not in df:
                subset = pd.DataFrame()
            else:
                subset = df[df["categories"].fillna("").str.contains(category, regex=False)].head(100)
            subset.to_excel(writer, index=False, sheet_name=category)
    _style_workbook(path)
